import os
import sys

from PyQt6 import QtCore, QtGui
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.checker import Checker, CheckerConfig, CheckerResult


def warn_user(parent, title: str, text: str) -> None:
    msg = QMessageBox(parent)
    msg.setIcon(QMessageBox.Icon.Warning)
    msg.setWindowTitle(title)
    msg.setText(text)
    msg.setStandardButtons(QMessageBox.StandardButton.Ok)
    msg.resize(500, 300)
    msg.exec()


def get_rel_path(data_path, slash_replace: bool = True) -> str:
    if getattr(sys, "frozen", False):
        try:
            base_path = sys._MEIPASS  # type: ignore[attr-defined]
            result = os.path.join(base_path, data_path)
        except Exception:
            return ""
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
        result = os.path.join(base_path, "..", "assets", data_path)

    if slash_replace:
        result = result.replace("\\", "/")

    return str(result)


def select_file(parent, title: str, filter_str: str) -> str:
    file_path, _ = QFileDialog.getOpenFileName(parent, title, "", filter_str)
    return file_path or ""


def select_entrance_table(self):
    file_path = select_file(self, "Выберите Excel файл", "Excel Files (*.xlsx *.xls)")
    if not file_path:
        return
    self.ui.lineEdit_entrances.setText(file_path)


def select_login_table(self):
    file_path = select_file(self, "Выберите Excel файл", "Excel Files (*.xlsx *.xls)")
    if not file_path:
        return
    self.ui.lineEdit_logins.setText(file_path)


def select_fired_list(self):
    file_path = select_file(self, "Выберите TXT файл", "Text Files (*.txt)")
    if not file_path:
        return
    self.ui.lineEdit_fireds.setText(file_path)


def select_exception_list(self):
    file_path = select_file(self, "Выберите TXT файл", "Text Files (*.txt);;Все файлы (*.*)", )
    if not file_path:
        return
    self.ui.lineEdit_exceptions.setText(file_path)


def inactive_ui(self):
    self.setCursor(QtCore.Qt.CursorShape.WaitCursor)
    self.ui.pushButton_check.setText("Проверка")
    self.ui.pushButton_check.setEnabled(False)
    self.ui.pushButton_check.setIcon(QtGui.QIcon(get_rel_path("loading.png")))
    self.ui.pushButton_check.setIconSize(QtCore.QSize(25, 25))
    self.ui.pushButton_save.setEnabled(False)


def active_ui(self):
    self.unsetCursor()
    self.ui.pushButton_check.setText("Начать поиск")
    self.ui.pushButton_check.setEnabled(True)
    self.ui.pushButton_check.setIcon(QtGui.QIcon(get_rel_path("start.png")))
    self.ui.pushButton_check.setIconSize(QtCore.QSize(25, 25))
    self.ui.pushButton_save.setEnabled(True)


def run_checker(self):
    entrances_table_path = (self.ui.lineEdit_entrances.text().strip().replace("/", "\\"))
    logins_table_path = self.ui.lineEdit_logins.text().strip().replace("/", "\\")
    exceptions_list_path = (self.ui.lineEdit_exceptions.text().strip().replace("/", "\\"))
    fired_list_path = self.ui.lineEdit_fireds.text().strip().replace("/", "\\")

    if not entrances_table_path:
        warn_user(self, "Заполните все поля", "Путь к журналу проходной пустой")
        return

    if not logins_table_path:
        warn_user(self, "Заполните все поля", "Путь к журналу входов в систему пустой", )
        return

    if entrances_table_path == logins_table_path:
        warn_user(self, "Неправильный таблицы", "Одинаковые пути к журналам")
        return

    if not os.path.isfile(entrances_table_path):
        warn_user(self, "Файл не найден", "Неправильный путь к журналу проходной")
        return

    if not os.path.isfile(logins_table_path):
        warn_user(self, "Файл не найден", "Неправильный путь к журналу входов в систему", )
        return

    if exceptions_list_path and not os.path.isfile(exceptions_list_path):
        warn_user(self, "Файл не найден", "Неправильный путь к списку исключений", )
        return

    if fired_list_path and not os.path.isfile(fired_list_path):
        warn_user(self, "Файл не найден", "Неправильный путь к списку уволенных", )
        return

    inactive_ui(self)

    checker = Checker(
        CheckerConfig(
            logins_table_path=logins_table_path,
            entrances_table_path=entrances_table_path,
            exceptions_list_path=exceptions_list_path,
            fired_list_path=fired_list_path,
            )
        )

    checker.signals.result.connect(lambda result: on_result(self, result))
    checker.signals.finished.connect(lambda: on_finished(self))
    checker.signals.error.connect(lambda e: on_error(self, e))
    QtCore.QThreadPool.globalInstance().start(checker)


def on_result(self, result: CheckerResult):
    self.scan_result = result

    self.ui.label_stats_leaks.setText(f"Утечек: {len(self.scan_result.compromised)}")
    self.ui.label_stats_leaks.setStyleSheet("color: #cc0000; font-weight:600")

    self.ui.label_stats_excluded.setText(f"Исключения: {len(self.scan_result.exceptions)}")
    self.ui.label_stats_excluded.setStyleSheet("font-weight:600")

    self.ui.label_stats_no_logins.setText(f"Нет входа: {len(self.scan_result.no_login)}")
    self.ui.label_stats_no_logins.setStyleSheet("color: #ffaa00; font-weight:600")

    self.ui.label_stats_matches.setText(f"Сопоставлено: {len(self.scan_result.matched)}")
    self.ui.label_stats_matches.setStyleSheet("color: #00c500; font-weight:600")

    total = (len(self.scan_result.compromised) + len(self.scan_result.exceptions) + len(self.scan_result.no_login) + len(self.scan_result.matched))
    self.ui.label_stats_total.setText(f"Всего: {total}")
    self.ui.label_stats_total.setStyleSheet("font-weight:600")


def on_finished(self):
    active_ui(self)


def on_error(self, error):
    self.scan_result = None
    warn_user(self, "Ошибка проверки", str(error))


def export_result(self):
    if self.scan_result is None:
        warn_user(self, "Нет результатов", "Нет результатов для выгрузки")
        return

    file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить файл как...", "result.xlsx", "Excel Files (*.xlsx *.xls);;Все файлы (*.*)", )

    if not file_path:
        return

    try:
        export_to_excel(self.scan_result, file_path)
    except Exception as e:
        warn_user(self, "Ошибка экспорта", str(e))


def export_to_excel(result: CheckerResult, path: str) -> None:
    categories = [
        ("Утечки", result.compromised),
        ("Исключения", result.exceptions),
        ("Нет входа", result.no_login),
        ("Сопоставлено", result.matched),
        ]

    headers = [
        "ФИО",
        "Отдел",
        "Таб. №",
        "User Name",
        "Client Host Name",
        "Logon Time",
        "Event Type Text",
        "Message",
        ]

    wb = Workbook()

    first_sheet_created = False

    for title, workers in categories:
        if not workers:
            continue

        if not first_sheet_created:
            ws = wb.active
            ws.title = title
            first_sheet_created = True
        else:
            ws = wb.create_sheet(title=title)

        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.showOutlineSymbols = True

        max_lengths: dict[int, int] = { }

        def update_max(_col_idx: int, value) -> None:
            if value is None:
                return
            length = len(str(value))
            prev = max_lengths.get(_col_idx, 0)
            if length > prev:
                max_lengths[_col_idx] = length

        # Заголовки
        ws.append(headers)
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            update_max(col_idx, cell.value)

        current_row = 2

        # Сортируем работников по ФИО
        workers.sort(key=lambda _worker: _worker.name)

        for worker in workers:
            # строка сотрудника
            ws.cell(row=current_row, column=1, value=worker.name)
            ws.cell(row=current_row, column=2, value=worker.department)
            ws.cell(row=current_row, column=3, value=worker.tab_number)

            update_max(1, worker.name)
            update_max(2, worker.department)
            update_max(3, worker.tab_number)

            current_row += 1

            if worker.logins:
                start_child_row = current_row

                for login in worker.logins:
                    ws.cell(row=current_row, column=4, value=login.user_name)
                    ws.cell(row=current_row, column=5, value=login.client_host_name)
                    ws.cell(row=current_row, column=6, value=login.logon_time)
                    ws.cell(row=current_row, column=7, value=login.event_type_text)
                    ws.cell(row=current_row, column=8, value=login.message)

                    update_max(4, login.user_name)
                    update_max(5, login.client_host_name)
                    update_max(6, login.logon_time)
                    update_max(7, login.event_type_text)
                    update_max(8, login.message)

                    current_row += 1

                # группируем строки входов
                ws.row_dimensions.group(start_child_row, current_row - 1, outline_level=1, hidden=True)

        # авто ширина на основе уже посчитанных max_lengths
        for col_idx, max_len in max_lengths.items():
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = min(70, max(12, max_len + 10))

    if not first_sheet_created:
        ws = wb.active
        ws.title = "Результаты"
        ws.append(["Нет данных для экспорта"])

    wb.save(path)
