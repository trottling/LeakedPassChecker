import os
import sys

from PyQt6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.checker import CheckerResult


def warn_user(parent, title: str, text: str) -> None:
    msg = QMessageBox(parent)
    msg.setIcon(QMessageBox.Icon.Warning)
    msg.setWindowTitle(title)
    msg.setText(text)
    msg.setStandardButtons(QMessageBox.StandardButton.Ok)
    msg.resize(500, 300)
    msg.exec()


def select_file(parent, title: str, filter_str: str) -> str:
    file_path, _ = QFileDialog.getOpenFileName(parent, title, "", filter_str)
    return file_path or ""


def get_rel_path(data_path, slash_replace: bool = True) -> str:
    if getattr(sys, "frozen", False):
        try:
            base_path = sys._MEIPASS  # type: ignore[attr-defined]
            result = os.path.join(base_path, data_path)
        except Exception:
            return ""
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
        result = os.path.join(base_path, "..", "assets", data_path)

    if slash_replace:
        result = result.replace("\\", "/")

    return str(result)


def export_to_excel(result: CheckerResult, path: str) -> None:
    categories = [
        ("Утечки", result.compromised),
        ("Исключения", result.exceptions),
        ("Нет входа", result.no_login),
        ("Сопоставлено", result.matched),
        ]

    headers = [
        "ФИО",
        "Отдел",
        "Таб. №",
        "User Name",
        "Client Host Name",
        "Logon Time",
        "Event Type Text",
        "Message",
        ]

    wb = Workbook()

    first_sheet_created = False

    for title, workers in categories:
        if not workers:
            continue

        if not first_sheet_created:
            ws = wb.active
            ws.title = title
            first_sheet_created = True
        else:
            ws = wb.create_sheet(title=title)

        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.showOutlineSymbols = True

        max_lengths: dict[int, int] = { }

        def update_max(_col_idx: int, value) -> None:
            if value is None:
                return
            length = len(str(value))
            prev = max_lengths.get(_col_idx, 0)
            if length > prev:
                max_lengths[_col_idx] = length

        # Заголовки
        ws.append(headers)
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            update_max(col_idx, cell.value)

        current_row = 2

        # Сортируем работников по ФИО
        workers.sort(key=lambda _worker: _worker.name)

        for worker in workers:
            # строка сотрудника
            ws.cell(row=current_row, column=1, value=worker.name)
            ws.cell(row=current_row, column=2, value=worker.department)
            ws.cell(row=current_row, column=3, value=worker.tab_number)

            update_max(1, worker.name)
            update_max(2, worker.department)
            update_max(3, worker.tab_number)

            current_row += 1

            if worker.logins:
                start_child_row = current_row

                for login in worker.logins:
                    ws.cell(row=current_row, column=4, value=login.user_name)
                    ws.cell(row=current_row, column=5, value=login.client_host_name)
                    ws.cell(row=current_row, column=6, value=login.logon_time)
                    ws.cell(row=current_row, column=7, value=login.event_type_text)
                    ws.cell(row=current_row, column=8, value=login.message)

                    update_max(4, login.user_name)
                    update_max(5, login.client_host_name)
                    update_max(6, login.logon_time)
                    update_max(7, login.event_type_text)
                    update_max(8, login.message)

                    current_row += 1

                # группируем строки входов
                ws.row_dimensions.group(start_child_row, current_row - 1, outline_level=1, hidden=True)

        # авто ширина на основе уже посчитанных max_lengths
        for col_idx, max_len in max_lengths.items():
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = min(70, max(12, max_len + 10))

    if not first_sheet_created:
        ws = wb.active
        ws.title = "Результаты"
        ws.append(["Нет данных для экспорта"])

    wb.save(path)
