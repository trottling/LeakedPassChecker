import re
from collections import defaultdict
from dataclasses import dataclass

from PyQt6.QtCore import QObject, QRunnable, pyqtSignal, pyqtSlot
from openpyxl import load_workbook

_LATIN_DIGITS_RE = re.compile(r"^[a-zA-Z0-9]+$")
_NON_CYRILLIC_RE = re.compile(r"[^А-Яа-я]+")
_YO_TRANSLATION = str.maketrans({ "Ё": "Е", "ё": "е" })


@dataclass(slots=True)
class Entrance:
    department: str
    fio: str
    tab_number: str


@dataclass(slots=True)
class Login:
    user_name: str
    client_host_name: str
    logon_time: str
    event_type_text: str
    failure_reason_text: str
    message: str
    user_display_name: str
    user_distinguish_name: str


@dataclass(slots=True)
class Worker:
    department: str
    name: str
    tab_number: str
    logins: list[Login]


@dataclass(slots=True)
class CheckerConfig:
    logins_table_path: str
    entrances_table_path: str
    exceptions_list_path: str
    fired_list_path: str


@dataclass(slots=True)
class CheckerResult:
    compromised: list[Worker]
    matched: list[Worker]
    no_login: list[Worker]
    exceptions: list[Worker]


class CheckerSignals(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(object)

    def __init__(self):
        super().__init__()


class Checker(QRunnable):
    LOGIN_HEADERS = [
        "user name",
        "client host name",
        "logon time",
        "event type text",
        "failure reason",
        "message",
        "user display name",
        "user distinguish name",
        ]

    ENTRANCE_HEADERS = [
        "дата",
        "отдел",
        "фио",
        "таб. №",
        "события",
        "события",
        ]

    def __init__(self, config: CheckerConfig):
        super().__init__()
        self.config = config
        self.exception_patterns: list[re.Pattern[str]] = []
        self.fired: set[str] = set()
        self.signals = CheckerSignals()

    @pyqtSlot()
    def run(self):
        try:
            result = self.execute()
            self.signals.result.emit(result)
        except Exception as e:
            self.signals.error.emit(str(e))
        finally:
            self.signals.finished.emit()

    def execute(self) -> CheckerResult:
        self.load_lists()

        entrances_by_fio = self.load_entrances_grouped()
        logins_by_fio = self.load_logins_grouped()

        compromised: list[Worker] = []
        exceptions: list[Worker] = []
        matched: list[Worker] = []
        no_login: list[Worker] = []

        for fio in entrances_by_fio.keys() | logins_by_fio.keys():
            fio_entrances = entrances_by_fio.get(fio, [])
            fio_logins = logins_by_fio.get(fio, [])

            first_entrance = fio_entrances[0] if fio_entrances else None

            worker = Worker(
                department=first_entrance.department if first_entrance and first_entrance.department else "-",
                name=fio,
                tab_number=first_entrance.tab_number if first_entrance and first_entrance.tab_number else "-",
                logins=fio_logins,
                )

            if fio in self.fired:
                exceptions.append(worker)
                continue

            has_entrance = bool(fio_entrances)
            has_login = bool(fio_logins)

            if has_login and has_entrance:
                matched.append(worker)
            elif has_login:
                if self.is_exception(worker):
                    exceptions.append(worker)
                else:
                    compromised.append(worker)
            else:
                no_login.append(worker)

        return CheckerResult(
            compromised=compromised,
            matched=matched,
            no_login=no_login,
            exceptions=exceptions,
            )

    def is_exception(self, worker: Worker) -> bool:
        for value in (worker.name, worker.department, worker.tab_number):
            if value and any(pattern.fullmatch(value) for pattern in self.exception_patterns):
                return True

        for login in worker.logins:
            for value in (
                    login.user_name,
                    login.client_host_name,
                    login.user_display_name,
                    login.user_distinguish_name,
                    ):
                if value and any(pattern.fullmatch(value) for pattern in self.exception_patterns):
                    return True

        return False

    def load_logins_grouped(self) -> dict[str, list[Login]]:
        wb = load_workbook(self.config.logins_table_path, data_only=True, read_only=True)
        try:
            ws = wb.active

            header_row = next(
                ws.iter_rows(min_row=11, max_row=11, min_col=1, max_col=8, values_only=True)
                )
            actual_headers = [normalize(v).lower() for v in header_row]

            if actual_headers != self.LOGIN_HEADERS:
                raise ValueError(
                    f"Неправильная таблица входов в систему\n\n"
                    f"Ожидаемые заголовки: {self.LOGIN_HEADERS}\n"
                    f"Найденные заголовки: {actual_headers}"
                    )

            result: dict[str, list[Login]] = defaultdict(list)

            for row in ws.iter_rows(min_row=12, max_col=8, values_only=True):
                if not any(row[:8]):
                    continue

                user_display_name = clean_fio(normalize(row[6]))
                user_distinguish_name = normalize(row[7])

                if not user_display_name:
                    user_display_name = fio_key(user_distinguish_name.split(",", 1)[0])
                else:
                    user_display_name = fio_key(user_display_name)

                login = Login(
                    user_name=normalize(row[0]),
                    client_host_name=normalize(row[1]),
                    logon_time=normalize(row[2]),
                    event_type_text=normalize(row[3]),
                    failure_reason_text=normalize(row[4]),
                    message=normalize(row[5]),
                    user_display_name=user_display_name,
                    user_distinguish_name=user_distinguish_name,
                    )

                result[user_display_name].append(login)

            return dict(result)
        finally:
            wb.close()

    def load_entrances_grouped(self) -> dict[str, list[Entrance]]:
        wb = load_workbook(self.config.entrances_table_path, data_only=True, read_only=True)
        try:
            ws = wb.active

            header_row = next(
                ws.iter_rows(min_row=7, max_row=7, min_col=1, max_col=6, values_only=True)
                )
            actual_headers = [normalize(v).lower() for v in header_row]

            if actual_headers != self.ENTRANCE_HEADERS:
                raise ValueError(
                    f"Неправильная таблица проходной\n\n"
                    f"Ожидаемые заголовки: {self.ENTRANCE_HEADERS}\n"
                    f"Найденные заголовки: {actual_headers}"
                    )

            result: dict[str, list[Entrance]] = defaultdict(list)

            for row in ws.iter_rows(min_row=9, max_col=6, values_only=True):
                department = normalize(row[1])
                raw_fio = normalize(row[2])
                tab_number = normalize(row[3])

                if not department and not raw_fio and not tab_number:
                    continue

                fio = fio_key(raw_fio)

                result[fio].append(
                    Entrance(
                        department=department,
                        fio=fio,
                        tab_number=tab_number,
                        )
                    )

            return dict(result)
        finally:
            wb.close()

    def load_lists(self) -> None:
        self.fired.clear()
        self.exception_patterns.clear()

        try:
            with open(self.config.fired_list_path, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    fio = clean_fio(line.strip())
                    if fio:
                        self.fired.add(fio)
        except FileNotFoundError:
            pass
        except Exception as e:
            raise ValueError(f"Ошибка чтения файла с ФИО уволенных: {e}")

        try:
            with open(self.config.exceptions_list_path, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    raw = line.strip()
                    if not raw:
                        continue
                    try:
                        self.exception_patterns.append(
                            re.compile(create_pattern(raw), re.IGNORECASE)
                            )
                    except re.error as e:
                        raise ValueError(f"Ошибка в паттерне исключения '{raw}': {e}")
        except FileNotFoundError:
            pass
        except Exception as e:
            raise ValueError(f"Ошибка чтения файла с исключениями: {e}")


def normalize(value) -> str:
    return "" if value is None else str(value).strip()


def clean_fio(text: str) -> str:
    text = normalize(text)
    if not text:
        return ""

    if _LATIN_DIGITS_RE.fullmatch(text):
        return text

    text = text.translate(_YO_TRANSLATION)
    text = _NON_CYRILLIC_RE.sub(" ", text).strip().lower()
    return text.title()


def fio_key(text: str) -> str:
    fio = clean_fio(text)
    return fio if len(fio) > 1 else "ОСТАЛЬНЫЕ"


def create_pattern(raw: str) -> str:
    pattern_re = re.escape(raw)
    pattern_re = pattern_re.replace(r"\*", ".*").replace(r"\?", ".")
    return pattern_re
