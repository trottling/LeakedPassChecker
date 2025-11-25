import os

from PyQt6 import QtCore, QtGui
from PyQt6.QtWidgets import QFileDialog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.checker import Checker, CheckerConfig
from app.utils import get_rel_path, warn_user


def select_entrance_table(self):
    try:
        file_path, _ = QFileDialog.getOpenFileName(None, "Выберите Excel файл", "", "Excel Files (*.xlsx *.xls)")

        if not file_path:
            return
    except:
        return

    self.ui.lineEdit_entrances.setText(file_path)


def select_login_table(self):
    try:
        file_path, _ = QFileDialog.getOpenFileName(None, "Выберите Excel файл", "", "Excel Files (*.xlsx *.xls)")

        if not file_path:
            return
    except:
        return

    self.ui.lineEdit_logins.setText(file_path)


def select_fired_list(self):
    try:
        file_path, _ = QFileDialog.getOpenFileName(None, "Выберите TXT файл", "", "Text Files (*.txt)")

        if not file_path:
            return
    except:
        return

    self.ui.lineEdit_fireds.setText(file_path)


def select_exception_table(self):
    try:
        file_path, _ = QFileDialog.getOpenFileName(None, "Выберите TXT файл с паттернами исключений", "", "Text Files (*.txt);;Все файлы (*.*)")

        if not file_path:
            return
    except:
        return

    self.ui.lineEdit_exceptions.setText(file_path)


def inactive_ui(self):
    self.ui.pushButton_check.setText("Проверка")
    self.ui.pushButton_check.setEnabled(False)
    self.ui.pushButton_check.setIcon(QtGui.QIcon(get_rel_path('loading.png')))
    self.ui.pushButton_check.setIconSize(QtCore.QSize(25, 25))
    self.ui.pushButton_save.setEnabled(False)


def active_ui(self):
    self.ui.pushButton_check.setText("Начать поиск")
    self.ui.pushButton_check.setEnabled(True)
    self.ui.pushButton_check.setIcon(QtGui.QIcon(get_rel_path('start.png')))
    self.ui.pushButton_check.setIconSize(QtCore.QSize(25, 25))
    self.ui.pushButton_save.setEnabled(True)


def run_checker(self):
    entrances_table_path = self.ui.lineEdit_entrances.text().strip()
    logins_table_path = self.ui.lineEdit_logins.text().strip()
    exceptions_table_path = self.ui.lineEdit_exceptions.text().strip()
    fired_list_path = self.ui.lineEdit_fireds.text().strip()

    if entrances_table_path == "":
        warn_user("Заполните все поля", "Путь к журналу проходной пустой")
        return

    if logins_table_path == "":
        warn_user("Заполните все поля", "Путь к журналу входов в систему пустой")
        return

    if entrances_table_path == logins_table_path:
        warn_user("Неправильный таблицы", "Одинаковые пути к журналам")
        return

    if not os.path.isfile(entrances_table_path):
        warn_user("Файл не найден", "Неправильный путь к журналу проходной")
        return

    if not os.path.isfile(logins_table_path):
        warn_user("Файл не найден", "Неправильный путь к журналу входов в систему")
        return

    if exceptions_table_path != "" and not os.path.isfile(exceptions_table_path):
        warn_user("Файл не найден", "Неправильный путь к файлу исключений")
        return

    if fired_list_path != "" and not os.path.isfile(fired_list_path):
        warn_user("Файл не найден", "Неправильный путь к списку уволенных")
        return

    inactive_ui(self)

    checker = Checker(
        CheckerConfig(
            logins_table_path=logins_table_path,
            entrances_table_path=entrances_table_path,
            exceptions_table_path=exceptions_table_path,
            fired_list_path=fired_list_path
            )
        )

    checker.signals.result.connect(lambda result: on_result(self, result))
    checker.signals.finished.connect(lambda: on_finished(self))
    checker.signals.error.connect(lambda e: on_error(self, e))
    checker.signals.stats.connect(lambda done, total: on_stat(self, done, total))
    QtCore.QThreadPool.globalInstance().start(checker)

def on_stat(self, done, total):
    self.ui.pushButton_check.setText(f"{done} / {total}")

def on_result(self, result):
    self.scan_result = result

    self.ui.label_stats_leaks.setText(f"Утечек: {self.scan_result.compromised_count}")
    self.ui.label_stats_leaks.setStyleSheet("color: #cc0000; font-weight:600")

    self.ui.label_stats_excluded.setText(f"Исключения: {self.scan_result.exceptions_count}")
    self.ui.label_stats_excluded.setStyleSheet("font-weight:600")

    self.ui.label_stats_no_logins.setText(f"Нет входа: {self.scan_result.no_login_count}")
    self.ui.label_stats_no_logins.setStyleSheet("color: #ffaa00; font-weight:600")

    self.ui.label_stats_matches.setText(f"Сопоставлено: {self.scan_result.matched_count}")
    self.ui.label_stats_matches.setStyleSheet("color: #00c500; font-weight:600")

    self.ui.label_stats_total.setText(f"Всего: {self.scan_result.total_count}")
    self.ui.label_stats_total.setStyleSheet("font-weight:600")


def on_finished(self):
    active_ui(self)


def on_error(self, error):
    self.scan_result = None
    warn_user("Ошибка проверки", str(error))


def export_result(self):
    if self.scan_result is None:
        warn_user("Нет результатов", "Нет результатов для выгрузки")
        return

    file_path, _ = QFileDialog.getSaveFileName(None, "Сохранить файл как...", "result.xlsx", "Excel Files (*.xlsx *.xls);;Все файлы (*.*)")

    if not file_path:
        return

    try:
        export_to_excel(self.scan_result, file_path)
    except Exception as e:
        warn_user("Ошибка экспорта", str(e))


def export_to_excel(result, path):
    categories = [
        ("Утечки", result.compromised),
        ("Исключения", result.exceptions),
        ("Нет входа", result.no_login),
        ("Сопоставлено", result.matched)
    ]
    headers = [
        "ФИО",
        "Отдел",
        "Таб. №",
        "User Name",
        "Client Host Name",
        "Logon Time",
        "Event Type Text",
        "Message"
    ]
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    def fill_sheet(ws, data):
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment
        current_row = 1

        for person in data:
            fio = person.get("fio") or person.get("user_display_name") or person.get("key")
            department = person.get("department") or ""
            tab_number = person.get("tab_number") or ""
            ws.append([fio, department, tab_number, "", "", "", "", ""])
            current_row += 1

            logins = person.get("logins") or []
            if logins:
                start_row = current_row + 1
                for login in logins:
                    ws.append([
                        "",
                        "",
                        "",
                        login.get("user_name"),
                        login.get("client_host_name"),
                        login.get("logon_time"),
                        login.get("event_type_text"),
                        login.get("message") or login.get("failure_reason")
                    ])
                    current_row += 1
                ws.row_dimensions.group(start_row, current_row, hidden=True)

        autosize(ws)

    def autosize(ws):
        for column_cells in ws.columns:
            column_letter = column_cells[0].column_letter
            max_length = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(70, max(12, max_length + 2))

    workbook = Workbook()
    first_sheet = True
    for name, data in categories:
        if first_sheet:
            sheet = workbook.active
            sheet.title = name
            first_sheet = False
        else:
            sheet = workbook.create_sheet(title=name)
        fill_sheet(sheet, data)

    workbook.save(path)