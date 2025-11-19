from PyQt6.QtCore import QObject, QRunnable, pyqtSignal, pyqtSlot
from openpyxl import load_workbook


class Checker_config:
    def __init__(self, logins_table_path: str, entrance_table_path: str, check_then_not_in_territory):
        self.check_then_not_in_territory = check_then_not_in_territory
        self.logins_table_path = logins_table_path
        self.entrance_table_path = entrance_table_path


class CheckerResult:
    def __init__(self, compromised, compromised_count, matched, matched_count, no_matches, no_matches_count):
        self.compromised = compromised
        self.compromised_count = compromised_count
        self.matched = matched
        self.matched_count = matched_count
        self.no_matches = no_matches
        self.no_matches_count = no_matches_count


class CheckerSignals(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(object)

    def __init__(self):
        super().__init__()


class Checker(QRunnable):
    def __init__(self, config: Checker_config):
        super().__init__()
        self.config = config
        self.signals = CheckerSignals()

    @pyqtSlot()
    def run(self):
        try:
            logins_wb = load_workbook(self.config.logins_table_path, data_only=True)
            entrance_wb = load_workbook(self.config.entrance_table_path, data_only=True)

            logins_ws = logins_wb.active
            entrance_ws = entrance_wb.active

            # Логины: ищем строку с "User Display Name"
            login_header_row_idx = None
            login_headers = None

            for i, row in enumerate(logins_ws.iter_rows(values_only=True), start=1):
                if not row:
                    continue
                if "User Display Name" in row:
                    login_header_row_idx = i
                    login_headers = list(row)
                    break

            if login_header_row_idx is None:
                raise ValueError("Не найден заголовок 'User Display Name' в таблице логинов")

            # Проходная: ищем строку с "ФИО"
            entrance_header_row_idx = None
            entrance_headers = None

            for i, row in enumerate(entrance_ws.iter_rows(values_only=True), start=1):
                if not row:
                    continue
                if "ФИО" in row:
                    entrance_header_row_idx = i
                    entrance_headers = list(row)
                    break

            if entrance_header_row_idx is None:
                raise ValueError("Не найден заголовок 'ФИО' в таблице проходной")

            # Читаем строки в словари

            def rows_to_dicts(ws, headers, start_row):
                records = []
                for row in ws.iter_rows(values_only=True, min_row=start_row):
                    if not row:
                        continue
                    # Приводим длину к длине headers (на случай пустых хвостов)
                    row = list(row)
                    if len(row) < len(headers):
                        row += [None] * (len(headers) - len(row))
                    record = { str(headers[i]): row[i] for i in range(len(headers)) }
                    records.append(record)
                return records

            logins_records = rows_to_dicts(
                logins_ws,
                login_headers,
                login_header_row_idx + 1,
                )
            entrance_records = rows_to_dicts(
                entrance_ws,
                entrance_headers,
                entrance_header_row_idx + 1,
                )

            # Группировка по имени

            logins_by_user = { }
            for rec in logins_records:
                name = rec.get("User Display Name")
                if not name:
                    continue
                key = str(name).strip()
                logins_by_user.setdefault(key, []).append(rec)

            entrance_by_user = { }
            for rec in entrance_records:
                fio = rec.get("ФИО")
                if not fio:
                    continue
                key = str(fio).strip()
                entrance_by_user.setdefault(key, []).append(rec)

            logins_users = set(logins_by_user.keys())
            entrance_users = set(entrance_by_user.keys())

            compromised_users = sorted(logins_users - entrance_users)
            matched_users = sorted(logins_users & entrance_users)
            no_match_users = sorted(entrance_users - logins_users)

            # Формируем результат

            compromised = []
            for user in compromised_users:
                compromised.append({
                    "user": user,
                    "logins_rows": logins_by_user.get(user, []),
                    "entrance_rows": [],  # нет записей на проходной
                    })

            matched = []
            for user in matched_users:
                matched.append({
                    "user": user,
                    "logins_rows": logins_by_user.get(user, []),
                    "entrance_rows": entrance_by_user.get(user, []),
                    })

            no_matches = []
            for user in no_match_users:
                no_matches.append({
                    "user": user,
                    "logins_rows": [],
                    "entrance_rows": entrance_by_user.get(user, []),
                    })

            self.signals.result.emit(CheckerResult(compromised, len(compromised_users), matched, len(matched_users), no_matches, len(no_match_users)))

        except Exception as e:
            self.signals.error.emit(str(e))
        finally:
            self.signals.finished.emit()
